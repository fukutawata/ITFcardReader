import openpyxl
import nfc
import re
import time
import os
import unicodedata
import jaconv
import pykakasi
import datetime

#ここにexcelの絶対パスを貼る
PASS = "-------.xlsx"

servc = 0x100B
number = 202011111
wb = openpyxl.load_workbook(PASS)
ws = wb["Sheet1"]
r = datetime.datetime.now()
c = 1
while not ws.cell(row=c, column=2).value is None:
    c = c + 1
def connected(tag):
    global number
    global c
    print(c)
    global ws
    service_code = [nfc.tag.tt3.ServiceCode(servc >> 6, servc & 0x3f)]
    bc_id = [nfc.tag.tt3.BlockCode(i) for i in range(3)]
    bc_name = [nfc.tag.tt3.BlockCode(3)]
    result=tag.read_without_encryption(service_code, bc_id)
    s=str(bytes(result))
    t=re.sub(r"\D", "", s)
    if number != t[2:11]:
        number = t[2:11]
        studentNumber = t[2:11]
        studentNameBeforeStripNull = tag.read_without_encryption(service_code, bc_name).decode('shift-jis')
        studentName = studentNameBeforeStripNull.strip().strip('\x00')
        studentMail = "s"+t[4:11]+"@u.tsukuba.ac.jp"
        os.system('play -n synth %s sin %s' % (0.2, 2000))
        
    
        print("学籍番号:",studentNumber) # 学籍番号
        print("氏名:",studentName) # 半角カナの名前
        print("mail:",studentMail)
        dt_now = datetime.datetime.now()
        ws.cell(row=c, column=1, value=str(dt_now))
        ws.cell(row=c, column=2, value=int(studentNumber))
        ws.cell(row=c, column=3, value=str(studentName))
        ws.cell(row=c, column=4, value=str(studentMail))
        wb.save("PASS") #上書き保存
        c = c + 1

        return False
 
# while True:
servc = 0x100B

clf = nfc.ContactlessFrontend('usb')
while True:
    try:
        clf.connect(rdwr={'on-connect': connected})
    except nfc.tag.tt3.Type3TagCommandError as e:
        print("aaaaaa")
        number = 202011111
    time.sleep(0.2)

wb.close()
